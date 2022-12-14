#Excel操作(读/写入)
"""
2003        .xls        xlrd        Excel三方库
2007/2008   .xlsx       openpyxl    Excel三方库
一、读取数据逻辑
1、先找到文件
2、打开文件
3、定位Excel里的sheet页面
4、遍历查找
5、数据返回
二、写入数据逻辑
1、先找到文件
2、打开文件
3、定位Excel里的sheet页面
4、确定总行数和写入的行数
5、往单元格中写入数据
6、保存数据(Excel)
7、
"""
class ExcelOpertions:
#     #读取Excel数据(xlrd)
#     def readExcelData(self, sheetname):
#         import xlrd
#         ExcelPath = '../poolData/user.xls'
#         OpenFile = xlrd.open_workbook(ExcelPath)
#         sheetpage = OpenFile.sheet_by_name(sheetname)
#         print(sheetpage.nrows)
#         userName = []
#         for i in range(1, sheetpage.nrows):
#             userName.append(sheetpage.cell_value(i, 0))
#         return userName

#读取Excel数据（xlrd）
    def readExcelData(self,sheetname,x,y):
        import xlrd
        ExcelPath='../poolData/user.xls'
        OpenFile=xlrd.open_workbook(ExcelPath)
        sheetpage=OpenFile.sheet_by_name(sheetname)
        print(sheetpage.nrows)
        userName=[]
        for i in range(1,sheetpage.nrows):
            userName.append(sheetpage.cell_value(i,0))
        return userName

# #个人练习
# class ExcelOpertions:
#     def readExcelData(self,userinfo,x,y):
#         import xlrd
#         ExcelPath='../poolData/user.xls'
#         OpenFile=xlrd.open_workbook(ExcelPath)
#         sheetpage=OpenFile.sheet_by_name('userinfo')
#         print(sheetpage.nrows)
#         userName=[]
#         for i in range(1, sheetpage.nrows):
#             userName.append(sheetpage.cell_value(i, 0))
#         return userName
# print(ExcelOpertions().readExcelData(,1,0))

#写入Excel数据(openpyxl)
#     def writeExcelData(self,write_data):
#         import openpyxl
#         #文件路径赋值给file
#         file="../poolData/user.xlsx"
#         #加载file(制作替身==可以理解为打开文件)，赋值给wb
#         wb=openpyxl.load_workbook(file)
#         #基于打开的文件，使用sheet页名称定位sheet页
#         sheetPage=wb['userinfo']
#         #打印当前Excel数据总行数
#         print(sheetPage.max_row)
#         #要写入数据的当前行数（写入数据的当前行数=当前Excel数据总行数+1），赋值给write_row
#         write_row=sheetPage.max_row+1
#         #基于定位的sheet页，往单元格（cell）写入数据，数据是活的使用变量write_data
#         sheetPage.cell(write_row,1,write_data)
#         #保存Excel
#         wb.save(file)
# # print(ExcelOpertions().readExcelData("userinfo",1,0))
# #Excel 操作类实例化，赋值给Excel
# Excel=ExcelOpertions()
# #调用写入Excel方法，并传入写入的数据
# print(Excel.writeExcelData("gebilaowang"))
    def writeExcelData(self,write_data):
        import openpyxl
        file="../poolData/user.xlsx"
        wb=openpyxl.load_workbook(file)
        sheetPage=wb["userinfo"]
        print(sheetPage.max_row)
        write_row=sheetPage.max_row+1
        sheetPage.cell(write_row,1,write_data)
        wb.save(file)
Excel=ExcelOpertions()
print(Excel.writeExcelData("wubaiwan"))




































